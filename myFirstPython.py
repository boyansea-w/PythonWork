# py -m pip install openpyxl
# py -m pip install tqdm
from openpyxl import load_workbook

def read_table_info(file_path, sheet_name='Sheet1'):
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    table_data = {}
    rowIdx = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过表头
        if all(cell is None for cell in row):
            continue  # 跳过空行

        table_data[rowIdx] = {
            "item_id" : row[1],
            "item_type" : row[2],
            "item_length" : row[3],
            "item_scal" : row[4],
            "item_isPkey" : row[5],
            "item_notnull" : row[6],
            "item_default" : row[7]
        }
        rowIdx += 1
    workbook.close()
    return table_data



import os

import time

start_time = time.time()
print("开始时间：", start_time)



# 设置目标文件夹路径
folder_path = r'D:\work\PythonWork'  # 替换为你自己的路径

# 获取所有 Excel 文件（包括 .xls 和 .xlsx）
excel_files = [f for f in os.listdir(folder_path) if f.endswith(('.xls', '.xlsx'))]

# 遍历并读取每个 Excel 文件
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    try:
        df = read_table_info(file_path)
        print(f"读取成功：{file}")
        for key in df[1].keys():
            print(f"{key}={df[1][key]}")
        df[1].keys()
        print(df.keys)  # 显示前几行数据
    except Exception as e:
        print(f"读取失败：{file}，错误信息：{e}")

end_time = time.time()
print("结束时间：", end_time)

print("运行时间：{:.2f} 秒".format(end_time - start_time))


